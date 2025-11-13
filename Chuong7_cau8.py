# doc_excel.py
from openpyxl import load_workbook

# Mở file Excel
wb = load_workbook('demo.xlsx')
print("Tên các sheet:", wb.sheetnames)

# Lấy sheet đầu tiên
ws = wb[wb.sheetnames[0]]

# Duyệt từng dòng và in ra các giá trị
for row in ws.values:
    for value in row:
        print(value, "\t", end='')
    print("")

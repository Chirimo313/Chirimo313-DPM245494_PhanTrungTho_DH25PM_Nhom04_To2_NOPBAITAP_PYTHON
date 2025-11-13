# quanlynhanvien_excel.py
import xlsxwriter
from openpyxl import load_workbook

path = "nhanvien.xlsx"

def luu_excel(ds):
    wb = xlsxwriter.Workbook(path)
    ws = wb.add_worksheet("NhanVien")
    ws.write_row('A1', ['Mã', 'Tên', 'Tuổi'])
    row = 2
    for nv in ds:
        ws.write(f"A{row}", nv[0])
        ws.write(f"B{row}", nv[1])
        ws.write(f"C{row}", nv[2])
        row += 1
    wb.close()

def doc_excel():
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    ds = []
    for i, row in enumerate(ws.values):
        if i == 0: continue
        ds.append(list(row))
    return ds

def them():
    ma = input("Mã NV: ")
    ten = input("Tên NV: ")
    tuoi = int(input("Tuổi: "))
    ds = doc_excel()
    ds.append([ma, ten, tuoi])
    luu_excel(ds)

def sapxep():
    ds = doc_excel()
    ds.sort(key=lambda x: x[2])
    luu_excel(ds)
    print("Đã sắp xếp theo tuổi tăng dần.")

while True:
    print("\n1.Thêm NV  2.Xem  3.Sắp xếp  0.Thoát")
    c = input("Chọn: ")
    if c == '1': them()
    elif c == '2': 
        for nv in doc_excel(): print(nv)
    elif c == '3': sapxep()
    elif c == '0': break
